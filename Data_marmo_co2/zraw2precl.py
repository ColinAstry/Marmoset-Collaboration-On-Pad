import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

def modify_xlsx_files(folder_path):
    # 创建名为 "precl" 的子文件夹
    output_folder = os.path.join(folder_path, 'precl')
    os.makedirs(output_folder, exist_ok=True)

    # 获取文件夹中的所有 XLSX 文件
    xlsx_files = [file for file in os.listdir(folder_path) if file.endswith('.xlsx')]

    for xlsx_file in xlsx_files:
        # 构建 XLSX 文件路径
        xlsx_path = os.path.join(folder_path, xlsx_file)

        # 读取 XLSX 文件
        wb = load_workbook(xlsx_path)
        ws = wb.active

        # #更改positionLeft positionTop positionLeftw positionTopw的值
        # # Define the mapping for the replacements
        # replacements = {0: 1, 0.33: 1, 0.66: 2}
        #  # Define the columns to replace
        # columns_to_replace = ['positionLeft', 'positionTop', 'positionLeftw', 'positionTopw']
        # # Replace the values in the specified columns
        # for column in columns_to_replace:
        #     combined_df[column] = combined_df[column].replace(replacements)

        # 获取 "time" 列的位置
        time_column = None
        for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
            for cell in col:
                if cell.value == 'time':
                    time_column = cell.column_letter
                    break
            if time_column:
                break

        if not time_column:
            print(f'在文件 {xlsx_file} 中找不到 "time" 列。跳过该文件。')
            continue

        # 插入 "times" 列和 "timem" 列
        time_column_index = column_index_from_string(time_column)
        ws.insert_cols(time_column_index + 1, 4)  # Inserting 4 columns instead of 2
        ws.cell(row=1, column=time_column_index + 1).value = 'times'
        ws.cell(row=1, column=time_column_index + 2).value = 'timem'
        ws.cell(row=1, column=time_column_index + 3).value = 't0'  # New column name
        ws.cell(row=1, column=time_column_index + 4).value = 't1'  # New column name

        for row in range(2, ws.max_row + 1):
            time_cell = ws.cell(row=row, column=time_column_index).value

            # 检查 "time" 列的值是否为 None
            if time_cell is None:
                print(f'在文件 {xlsx_file} 第 {row} 行的 "time" 列中存在空值。跳过该行。')
                continue

            # 进行后续处理
            time_cell = int(time_cell)  # 将 time_cell 转换为整数类型
            times_value = time_cell / 1000
            timem_value = time_cell / 60000

            ws.cell(row=row, column=time_column_index + 1).value = times_value
            ws.cell(row=row, column=time_column_index + 2).value = timem_value
            ws.cell(row=row, column=time_column_index + 3).value = ''  # Leave it empty for "t0"
            ws.cell(row=row, column=time_column_index + 4).value = ''  # Leave it empty for "t1"
    
        # 获取最后一列的索引
        last_column_index = ws.max_column

        # 插入 "note" 列
        note_column_index = last_column_index + 1  # Insert "note" column at the end of the chart
        ws.insert_cols(note_column_index)
        ws.cell(row=1, column=note_column_index).value = 'note'

        # Process the rows
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=note_column_index).value = ''  # Leave the "note" column empty

        # 保存 XLSX 文件
        output_file = os.path.join(output_folder, xlsx_file)
        wb.save(output_file)

        print(f'修改完成：{output_file}')

    print('所有 XLSX 文件已修改。')


import sys

# The first argument is the script name itself, so we get the second argument
# folder_path = sys.argv[1]  # replace with your actual folder path
folder_path = r'D:\Research\MarmoCo2\Data\ToFilter'
modify_xlsx_files(folder_path)